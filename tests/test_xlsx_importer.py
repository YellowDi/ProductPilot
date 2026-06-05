from __future__ import annotations

import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from importlib.util import find_spec
from io import StringIO
from pathlib import Path

if find_spec("openpyxl") is not None:
    from openpyxl import Workbook
else:
    Workbook = None  # type: ignore[assignment]

from product_pilot.cli import validate_product
from product_pilot.importers.xlsx import load_products_from_xlsx


@unittest.skipUnless(Workbook is not None, "openpyxl is not installed")
class ProductXlsxImporterTests(unittest.TestCase):
    def test_loads_products_from_chinese_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "main.jpg").write_bytes(b"test")
            (root / "detail-01.jpg").write_bytes(b"test")
            (root / "sku.jpg").write_bytes(b"test")
            workbook_path = root / "product-input.xlsx"
            _write_product_workbook(workbook_path)

            products = load_products_from_xlsx(workbook_path)

            self.assertEqual(len(products), 1)
            product = products[0]
            self.assertEqual(product.product_id, "SKU06")
            self.assertEqual(product.title, "UNITY优妮蒂男鞋低帮板鞋舒适休闲鞋")
            self.assertEqual(product.category, "流行男鞋 > 低帮鞋 > 板鞋")
            self.assertEqual([sku.name for sku in product.skus], ["41", "42"])
            self.assertEqual([image.role for image in product.images], ["main", "detail", "sku"])

    def test_validate_product_accepts_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "main.jpg").write_bytes(b"test")
            (root / "detail-01.jpg").write_bytes(b"test")
            (root / "sku.jpg").write_bytes(b"test")
            workbook_path = root / "product-input.xlsx"
            _write_product_workbook(workbook_path)

            with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
                self.assertEqual(validate_product(workbook_path), 0)


def _write_product_workbook(path: Path) -> None:
    assert Workbook is not None
    workbook = Workbook()
    products = workbook.active
    products.title = "商品"
    products.append(["商品编号", "商品标题", "类目路径", "商品描述"])
    products.append(["SKU06", "UNITY优妮蒂男鞋低帮板鞋舒适休闲鞋", "流行男鞋 > 低帮鞋 > 板鞋", "fixture"])

    skus = workbook.create_sheet("SKU")
    skus.append(["商品编号", "SKU名称", "拼单价", "单买价", "参考价", "库存"])
    skus.append(["SKU06", "41", "29.90", "39.90", "99.00", 8])
    skus.append(["SKU06", "42", "29.90", "39.90", "99.00", 10])

    images = workbook.create_sheet("图片")
    images.append(["商品编号", "图片角色", "图片文件名"])
    images.append(["SKU06", "main", "main.jpg"])
    images.append(["SKU06", "detail", "detail-01.jpg"])
    images.append(["SKU06", "sku", "sku.jpg"])

    workbook.save(path)


if __name__ == "__main__":
    unittest.main()
