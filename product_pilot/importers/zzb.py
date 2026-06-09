"""Importer for Zhizunbao ecommerce plugin exports."""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from product_pilot.domain.product import ProductDraft


class ZzbImportError(ValueError):
    pass


@dataclass(frozen=True)
class ZzbSkuTextRow:
    color: str
    size: str
    price: Decimal


@dataclass(frozen=True)
class ZzbImportRequest:
    excel_path: Path
    sku_text: str
    assets_path: Path
    title: str
    category: str
    output_path: Path
    product_id: str = ""
    product_code: str = ""
    stock: int | None = None
    group_price: Decimal | None = None
    single_price: Decimal | None = None


@dataclass(frozen=True)
class ZzbImportResult:
    output_path: Path
    product: ProductDraft
    asset_root: Path
    notes: tuple[str, ...] = ()


@dataclass(frozen=True)
class _ZzbExcelRow:
    source: str
    name: str
    price: Decimal | None = None
    stock: int | None = None


_SKU_TEXT_PATTERN = re.compile(
    r"颜色分类\s*[:：]\s*(?P<color>.+?)\s+"
    r"鞋码\s*[:：]\s*(?P<size>\S+)"
    r".*?"
    r"价格\s*[:：]\s*(?P<price>\d+(?:\.\d+)?)"
)
_SKU_FILENAME_PATTERN = re.compile(r"^SKU\d+_(?P<color>[^_]+)_(?P<size>\S+)")
_PRODUCT_ID_PATTERN = re.compile(r"商品ID[_-]?(\d+)")


def import_zzb_export(request: ZzbImportRequest) -> ZzbImportResult:
    title = request.title.strip()
    category = request.category.strip()
    if not title:
        raise ZzbImportError("商品标题不能为空")
    if not category:
        raise ZzbImportError("类目路径不能为空")

    output_path = request.output_path.expanduser()
    asset_root, notes = _prepare_asset_root(request.assets_path.expanduser(), output_path.parent)
    rows = _read_zzb_excel_rows(request.excel_path.expanduser())
    sku_text_rows = parse_zzb_sku_text(request.sku_text)
    if not sku_text_rows:
        raise ZzbImportError("SKU文字未解析出任何规格")

    product_id = request.product_id.strip() or _product_id_from_path(request.assets_path) or output_path.parent.name
    product_code = request.product_code.strip() or product_id
    if request.stock is not None and request.stock < 0:
        raise ZzbImportError("库存不能小于 0")
    if request.group_price is not None and request.group_price <= 0:
        raise ZzbImportError("拼单价必须大于 0")
    if request.single_price is not None and request.single_price <= 0:
        raise ZzbImportError("单买价必须大于 0")
    if (
        request.group_price is not None
        and request.single_price is not None
        and request.single_price <= request.group_price
    ):
        raise ZzbImportError("单买价必须大于拼单价")
    product = _build_product(
        rows=rows,
        sku_text_rows=sku_text_rows,
        asset_root=asset_root,
        workbook_dir=output_path.parent,
        product_id=product_id,
        product_code=product_code,
        title=title,
        category=category,
        stock=request.stock,
        group_price=request.group_price,
        single_price=request.single_price,
    )
    errors = product.validate()
    if errors:
        raise ZzbImportError("\n".join(errors))

    _write_standard_workbook(output_path, product)
    return ZzbImportResult(
        output_path=output_path,
        product=product,
        asset_root=asset_root,
        notes=tuple(notes),
    )


def suggest_zzb_output_path(base_dir: Path, excel_path: Path, assets_path: Path) -> Path:
    product_id = _product_id_from_path(assets_path)
    name = product_id or assets_path.expanduser().stem or excel_path.expanduser().stem
    return base_dir.expanduser() / _safe_path_segment(name) / "product-input.xlsx"


def parse_zzb_sku_text(text: str) -> list[ZzbSkuTextRow]:
    rows: list[ZzbSkuTextRow] = []
    errors: list[str] = []
    for line_number, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        match = _SKU_TEXT_PATTERN.search(line)
        if match is None:
            errors.append(f"SKU文字第 {line_number} 行无法解析：{line}")
            continue
        price = _decimal(match.group("price"))
        if price is None or price <= 0:
            errors.append(f"SKU文字第 {line_number} 行价格无效：{line}")
            continue
        rows.append(
            ZzbSkuTextRow(
                color=match.group("color").strip(),
                size=match.group("size").strip(),
                price=price,
            )
        )

    if errors:
        raise ZzbImportError("\n".join(errors))
    return rows


def _prepare_asset_root(assets_path: Path, output_dir: Path) -> tuple[Path, list[str]]:
    if not assets_path.exists():
        raise ZzbImportError(f"媒体资源不存在：{assets_path}")

    if assets_path.is_dir():
        return _find_asset_root(assets_path), []

    if assets_path.suffix.lower() != ".zip":
        raise ZzbImportError("媒体资源必须是 zip 文件或已解压目录")

    extract_dir = output_dir / "media"
    extract_dir.mkdir(parents=True, exist_ok=True)
    try:
        with ZipFile(assets_path) as zip_file:
            for member in zip_file.infolist():
                filename = member.filename
                if not filename or filename.startswith("__MACOSX/") or filename.endswith(".DS_Store"):
                    continue
                target = (extract_dir / filename).resolve()
                if not target.is_relative_to(extract_dir.resolve()):
                    raise ZzbImportError(f"zip 包含不安全路径：{filename}")
                zip_file.extract(member, extract_dir)
    except BadZipFile as exc:
        raise ZzbImportError(f"zip 文件无效：{assets_path}") from exc

    return _find_asset_root(extract_dir), [f"已解压媒体 zip：{extract_dir.resolve()}"]


def _find_asset_root(path: Path) -> Path:
    candidates = [path, *[item for item in path.iterdir() if item.is_dir()]]
    for candidate in candidates:
        if (candidate / "主图").is_dir() and (candidate / "详情图").is_dir() and (candidate / "SKU").is_dir():
            return candidate
    raise ZzbImportError(f"媒体目录缺少 主图/详情图/SKU 子目录：{path}")


def _read_zzb_excel_rows(path: Path) -> list[_ZzbExcelRow]:
    if not path.exists():
        raise ZzbImportError(f"至尊宝表格不存在：{path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ZzbImportError("读取 XLSX 需要 openpyxl，请先安装项目依赖") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ZzbImportError(f"至尊宝表格无效：{exc}") from exc

    try:
        sheet = workbook["全部"] if "全部" in workbook.sheetnames else workbook.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise ZzbImportError("至尊宝表格为空")
        indexes = {str(value or "").strip(): index for index, value in enumerate(header)}
        required = ("来源", "名称")
        missing = [name for name in required if name not in indexes]
        if missing:
            raise ZzbImportError(f"至尊宝表格缺少列：{', '.join(missing)}")

        rows: list[_ZzbExcelRow] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            source = _cell_text(values, indexes["来源"])
            name = _cell_text(values, indexes["名称"])
            if not source and not name:
                continue
            if not source or not name:
                raise ZzbImportError(f"至尊宝表格存在来源或名称为空的行：{values}")
            rows.append(
                _ZzbExcelRow(
                    source=source,
                    name=name,
                    price=_decimal(_cell_value(values, indexes.get("价格"))),
                    stock=_int(_cell_value(values, indexes.get("库存"))),
                )
            )

        if not rows:
            raise ZzbImportError("至尊宝表格没有有效数据行")
        return rows
    finally:
        workbook.close()


def _build_product(
    *,
    rows: list[_ZzbExcelRow],
    sku_text_rows: list[ZzbSkuTextRow],
    asset_root: Path,
    workbook_dir: Path,
    product_id: str,
    product_code: str,
    title: str,
    category: str,
    stock: int | None,
    group_price: Decimal | None,
    single_price: Decimal | None,
) -> ProductDraft:
    sku_excel_rows = [row for row in rows if row.source == "SKU"]
    stock_by_sku = _stock_by_sku(sku_excel_rows)
    sku_file_by_color = _sku_file_by_color(sku_excel_rows, asset_root)

    skus = []
    for row in sku_text_rows:
        key = (row.color, row.size)
        sku_price = group_price if group_price is not None else row.price
        sku_single_price = single_price if single_price is not None else sku_price + Decimal("1.00")
        if sku_single_price <= sku_price:
            raise ZzbImportError("单买价必须大于拼单价")
        skus.append(
            {
                "name": f"{row.color} {row.size}",
                "attributes": {"颜色分类": row.color, "鞋码": row.size},
                "price": sku_price,
                "single_price": sku_single_price,
                "stock": stock if stock is not None else stock_by_sku.get(key, 1000),
            }
        )

    images = []
    for row in rows:
        if row.source == "主图":
            images.append({"path": _workbook_image_path(workbook_dir, asset_root / "主图" / row.name), "role": "main"})
        elif row.source == "详情图":
            images.append({"path": _workbook_image_path(workbook_dir, asset_root / "详情图" / row.name), "role": "detail"})

    seen_colors: set[str] = set()
    for sku in sku_text_rows:
        if sku.color in seen_colors:
            continue
        image_path = sku_file_by_color.get(sku.color)
        if image_path is None:
            raise ZzbImportError(f"缺少颜色 SKU 图片：{sku.color}")
        seen_colors.add(sku.color)
        images.append(
            {
                "path": _workbook_image_path(workbook_dir, image_path),
                "role": "sku",
                "sku_attribute": "颜色分类",
                "sku_value": sku.color,
            }
        )

    _ensure_declared_images_exist(images, workbook_dir)
    return ProductDraft.from_mapping(
        {
            "product_id": product_id,
            "product_code": product_code,
            "title": title,
            "category": category,
            "images": images,
            "skus": skus,
        }
    )


def _stock_by_sku(rows: list[_ZzbExcelRow]) -> dict[tuple[str, str], int]:
    stocks: dict[tuple[str, str], int] = {}
    for row in rows:
        parsed = _parse_sku_filename(row.name)
        if parsed is None or row.stock is None:
            continue
        stocks.setdefault(parsed, row.stock)
    return stocks


def _sku_file_by_color(rows: list[_ZzbExcelRow], asset_root: Path) -> dict[str, Path]:
    images: dict[str, Path] = {}
    for row in rows:
        parsed = _parse_sku_filename(row.name)
        if parsed is None:
            continue
        color, _ = parsed
        path = asset_root / "SKU" / row.name
        if color not in images and path.exists():
            images[color] = path
    return images


def _parse_sku_filename(name: str) -> tuple[str, str] | None:
    match = _SKU_FILENAME_PATTERN.search(Path(name).stem)
    if match is None:
        return None
    return match.group("color").strip(), match.group("size").strip()


def _write_standard_workbook(path: Path, product: ProductDraft) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ZzbImportError("写入 XLSX 需要 openpyxl，请先安装项目依赖") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    products = workbook.active
    products.title = "商品"
    products.append(["商品编号", "商品标题", "类目路径", "商品描述", "默认拼单价", "默认单买价", "默认参考价", "默认库存", "商品编码"])
    products.append([product.product_id, product.title, product.category, product.description, "", "", "", "", product.product_code])

    skus = workbook.create_sheet("SKU")
    skus.append(["商品编号", "颜色分类", "鞋码", "SKU名称", "拼单价", "单买价", "参考价", "库存"])
    for sku in product.skus:
        skus.append(
            [
                product.product_id,
                sku.attributes.get("颜色分类", ""),
                sku.attributes.get("鞋码", ""),
                "",
                str(sku.price),
                str(sku.single_price or ""),
                str(sku.reference_price or ""),
                sku.stock,
            ]
        )

    images = workbook.create_sheet("图片")
    images.append(["商品编号", "图片角色", "图片文件名", "SKU属性", "SKU值"])
    for image in product.images:
        images.append([product.product_id, image.role, str(image.path), image.sku_attribute, image.sku_value])

    try:
        workbook.save(path)
    finally:
        workbook.close()


def _ensure_declared_images_exist(images: list[dict[str, Any]], workbook_dir: Path) -> None:
    missing = []
    for image in images:
        path = Path(str(image["path"]))
        full_path = path if path.is_absolute() else workbook_dir / path
        if not full_path.is_file():
            missing.append(str(full_path))
    if missing:
        raise ZzbImportError("媒体文件不存在：\n" + "\n".join(missing[:10]))


def _workbook_image_path(workbook_dir: Path, image_path: Path) -> str:
    resolved = image_path.resolve()
    try:
        return str(resolved.relative_to(workbook_dir.resolve()))
    except ValueError:
        return str(resolved)


def _product_id_from_path(path: Path) -> str:
    match = _PRODUCT_ID_PATTERN.search(path.stem)
    return match.group(1) if match else ""


def _safe_path_segment(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z._-]+", "-", value.strip())
    return cleaned.strip("-") or "zzb-product"


def _cell_text(values: tuple[Any, ...], index: int) -> str:
    return str(_cell_value(values, index) or "").strip()


def _cell_value(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def _decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None
