"""XLSX product workbook importer."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from product_pilot.domain.product import ProductDraft


class ProductWorkbookError(ValueError):
    pass


DEFAULT_XLSX_STOCK = 1000


SHEET_ALIASES = {
    "products": ("商品", "Products"),
    "skus": ("SKU", "SKUs"),
    "images": ("图片", "Images"),
}

HEADER_ALIASES = {
    "product_id": ("商品编号", "product_id"),
    "title": ("商品标题", "title"),
    "category": ("类目路径", "category"),
    "description": ("商品描述", "description"),
    "default_price": ("默认拼单价", "default_price"),
    "default_single_price": ("默认单买价", "default_single_price"),
    "default_reference_price": ("默认参考价", "default_reference_price"),
    "default_stock": ("默认库存", "default_stock"),
    "sku_name": ("SKU名称", "sku_name", "name"),
    "color": ("颜色分类", "color"),
    "size": ("鞋码", "尺码", "size"),
    "price": ("拼单价", "price"),
    "single_price": ("单买价", "single_price"),
    "reference_price": ("参考价", "reference_price"),
    "stock": ("库存", "stock"),
    "image_role": ("图片角色", "image_role", "role"),
    "image_path": ("图片文件名", "图片路径", "文件名", "image_path", "path"),
    "image_sku_attribute": ("SKU属性", "绑定SKU属性", "sku_attribute"),
    "image_sku_value": ("SKU值", "绑定SKU值", "sku_value"),
}


def load_products_from_xlsx(path: Path) -> list[ProductDraft]:
    if not path.exists():
        raise ProductWorkbookError(f"file not found: {path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ProductWorkbookError("XLSX support requires openpyxl. Install project dependencies first.") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ProductWorkbookError(f"invalid xlsx: {exc}") from exc

    products_sheet = _get_sheet(workbook, "products")
    skus_sheet = _get_sheet(workbook, "skus")
    images_sheet = _get_sheet(workbook, "images")

    products = _read_products(products_sheet)
    skus_by_product_id = _read_skus(skus_sheet, products)
    images_by_product_id = _read_images(images_sheet, set(products))

    return [
        ProductDraft.from_mapping(
            {
                **payload,
                "skus": skus_by_product_id.get(product_id, []),
                "images": images_by_product_id.get(product_id, []),
            }
        )
        for product_id, payload in products.items()
    ]


def _get_sheet(workbook: Any, key: str) -> Any:
    for name in SHEET_ALIASES[key]:
        if name in workbook.sheetnames:
            return workbook[name]
    expected = ", ".join(SHEET_ALIASES[key])
    raise ProductWorkbookError(f"missing sheet: {expected}")


def _read_products(sheet: Any) -> dict[str, dict[str, Any]]:
    rows = _read_rows(
        sheet,
        required=("product_id", "title", "category"),
        optional=(
            "description",
            "default_price",
            "default_single_price",
            "default_reference_price",
            "default_stock",
        ),
    )
    products: dict[str, dict[str, Any]] = {}
    errors: list[str] = []
    for row_number, row in rows:
        product_id = _text(row.get("product_id"))
        if not product_id:
            errors.append(f"{sheet.title} row {row_number}: 商品编号 is required")
            continue
        if product_id in products:
            errors.append(f"{sheet.title} row {row_number}: duplicate 商品编号: {product_id}")
            continue
        products[product_id] = {
            "product_id": product_id,
            "title": _text(row.get("title")),
            "category": _text(row.get("category")),
            "description": _text(row.get("description")),
            "default_price": row.get("default_price"),
            "default_single_price": row.get("default_single_price"),
            "default_reference_price": row.get("default_reference_price"),
            "default_stock": row.get("default_stock"),
        }

    if errors:
        raise ProductWorkbookError("\n".join(errors))
    if not products:
        raise ProductWorkbookError(f"{sheet.title} sheet has no product rows")
    return products


def _read_skus(sheet: Any, products: dict[str, dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    rows = _read_rows(
        sheet,
        required=("product_id",),
        optional=("sku_name", "color", "size", "price", "single_price", "reference_price", "stock"),
    )
    product_ids = set(products)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    errors: list[str] = []
    for row_number, row in rows:
        product_id = _resolve_related_product_id(row.get("product_id"), products)
        if product_id is None:
            errors.append(f"{sheet.title} row {row_number}: 商品编号 is required")
            continue
        if product_id not in product_ids:
            errors.append(f"{sheet.title} row {row_number}: unknown 商品编号: {product_id}")
            continue
        attributes = _sku_attributes(row)
        name = _text(row.get("sku_name")) or _sku_name_from_attributes(attributes)
        if not name:
            errors.append(f"{sheet.title} row {row_number}: SKU名称 or at least one SKU attribute is required")
            continue
        product = products[product_id]
        grouped[product_id].append(
            {
                "name": name,
                "attributes": attributes,
                "price": _value_or_default(row.get("price"), product.get("default_price")),
                "single_price": _value_or_default(row.get("single_price"), product.get("default_single_price")),
                "reference_price": _value_or_default(
                    row.get("reference_price"),
                    product.get("default_reference_price"),
                ),
                "stock": _value_or_default(row.get("stock"), product.get("default_stock"), DEFAULT_XLSX_STOCK),
            }
        )

    if errors:
        raise ProductWorkbookError("\n".join(errors))
    return grouped


def _read_images(sheet: Any, product_ids: set[str]) -> dict[str, list[dict[str, Any]]]:
    rows = _read_rows(
        sheet,
        required=("product_id", "image_role", "image_path"),
        optional=("image_sku_attribute", "image_sku_value"),
    )
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    products = {product_id: {} for product_id in product_ids}
    errors: list[str] = []
    for row_number, row in rows:
        product_id = _resolve_related_product_id(row.get("product_id"), products)
        if product_id is None:
            errors.append(f"{sheet.title} row {row_number}: 商品编号 is required")
            continue
        if product_id not in product_ids:
            errors.append(f"{sheet.title} row {row_number}: unknown 商品编号: {product_id}")
            continue
        grouped[product_id].append(
            {
                "path": _text(row.get("image_path")),
                "role": _text(row.get("image_role")) or "gallery",
                "sku_attribute": _text(row.get("image_sku_attribute")),
                "sku_value": _text(row.get("image_sku_value")),
            }
        )

    if errors:
        raise ProductWorkbookError("\n".join(errors))
    return grouped


def _read_rows(sheet: Any, *, required: Iterable[str], optional: Iterable[str]) -> list[tuple[int, dict[str, Any]]]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ProductWorkbookError(f"{sheet.title} sheet is empty")

    header_indexes = _header_indexes(header_row)
    missing_headers = [
        key
        for key in required
        if not any(alias in header_indexes for alias in _normalized_aliases(key))
    ]
    if missing_headers:
        labels = ", ".join(HEADER_ALIASES[key][0] for key in missing_headers)
        raise ProductWorkbookError(f"{sheet.title} sheet missing required columns: {labels}")

    selected_keys = tuple(required) + tuple(optional)
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(values):
            continue
        rows.append((row_number, _map_row(values, header_indexes, selected_keys)))
    return rows


def _header_indexes(header_row: tuple[Any, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = _normalize(value)
        if header:
            indexes[header] = index
    return indexes


def _map_row(values: tuple[Any, ...], header_indexes: dict[str, int], keys: tuple[str, ...]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for key in keys:
        for alias in _normalized_aliases(key):
            if alias in header_indexes:
                index = header_indexes[alias]
                mapped[key] = values[index] if index < len(values) else None
                break
    return mapped


def _normalized_aliases(key: str) -> tuple[str, ...]:
    return tuple(_normalize(alias) for alias in HEADER_ALIASES[key])


def _normalize(value: Any) -> str:
    return str(value or "").strip().lower()


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_row(values: tuple[Any, ...]) -> bool:
    return all(_text(value) == "" for value in values)


def _resolve_related_product_id(value: Any, products: dict[str, dict[str, Any]]) -> str | None:
    product_id = _text(value)
    if product_id:
        return product_id
    if len(products) == 1:
        return next(iter(products))
    return None


def _sku_attributes(row: dict[str, Any]) -> dict[str, str]:
    attributes: dict[str, str] = {}
    color = _text(row.get("color"))
    size = _text(row.get("size"))
    if color:
        attributes["颜色分类"] = color
    if size:
        attributes["鞋码"] = size
    return attributes


def _sku_name_from_attributes(attributes: dict[str, str]) -> str:
    return " ".join(attributes.values())


def _value_or_default(value: Any, default: Any, fallback: Any = "") -> Any:
    if _text(value) != "":
        return value
    if _text(default) != "":
        return default
    return fallback
